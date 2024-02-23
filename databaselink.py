from openpyxl import Workbook, load_workbook

with open('dateno.txt', 'r') as no:
    num = no.readline()
    filename = f"data{num}.xlsx"

    def write_to_excel(No, Rank, Name, Price, DateUpdate, TimeUpdate):
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["No", "Rank", "Name", "Price", "Date-Updated", "Time-Updated"])

        row = [No, Rank, Name, Price, DateUpdate, TimeUpdate]
        ws.append(row)
        wb.save(filename)


    num = int(num) + 1
    with open('dateno.txt', 'w') as no1:
        no1.write(str(num))
    #    def write_to_excel(No, Rank, Name, Price, DateUpdate, TimeUpdate):
